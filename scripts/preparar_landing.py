"""Extrae hojas del xlsx de mediciones en formato ANCHO para dbt/macro unpivot.

Formato de salida (lo que espera unpivot_mediciones_planta):
  - Una fila por atributo (etiqueta, Semanas, Fecha, Operador, punto1, punto2, ...)
  - Columnas: etiqueta, column01, column02, ..., column71

Uso:
    python scripts/preparar_landing.py --raiz /cliente/minera
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

HOJAS: dict[str, str] = {
    "Prechancado":             "Prechancado.csv",
    "2° y 3°":                 "2_y_3.csv",
    "Terciario":               "Terciario.csv",
    "Cuaternario":             "Cuaternario.csv",
    "Molienda Sag":            "Molienda_Sag.csv",
    "Molienda Convencional":   "Molienda_Convencional.csv",
    "CDM (1)":                 "CDM_(1).csv",
    "CDM (2)":                 "CDM_(2).csv",
    "Nodo":                    "Nodo.csv",
}

FILA_PRIMER_DATO = 2   # Fila 1 es título del proceso; desde fila 2 hay datos útiles


def serializar(valor: Any) -> str:
    if valor is None:
        return ""
    from datetime import date, datetime
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    if isinstance(valor, float):
        return str(int(valor)) if valor.is_integer() else f"{valor:.6f}".rstrip("0").rstrip(".")
    return str(valor)


def extraer_hoja_ancho(ws: Any) -> tuple[list[str], list[list[str]]]:
    """Extrae la hoja en formato ancho: una fila por atributo, una columna por semana.

    Determina el número de columnas de datos contando valores en la fila 2
    (fila de semanas). Retorna (encabezados, filas).
    Encabezados: ['etiqueta', 'column01', 'column02', ..., 'columnNN']
    Filas: cada fila es [etiqueta, valor_col1, valor_col2, ...]
    """
    # Detectar número de columnas de datos (fila 2 = "Semanas")
    n_cols = 0
    for cell in ws[FILA_PRIMER_DATO][1:]:
        if cell.value is None:
            break
        n_cols += 1

    if n_cols == 0:
        return [], []

    # Encabezados: etiqueta + column01..columnNN
    encabezados = ["etiqueta"] + [f"column{str(i+1).zfill(2)}" for i in range(n_cols)]

    filas: list[list[str]] = []
    for r in range(FILA_PRIMER_DATO, ws.max_row + 1):
        etiqueta = ws.cell(row=r, column=1).value
        if etiqueta is None:
            continue
        fila = [serializar(etiqueta)]
        for c in range(2, n_cols + 2):
            fila.append(serializar(ws.cell(row=r, column=c).value))
        filas.append(fila)

    return encabezados, filas


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--raiz", type=Path, default=Path.cwd())
    args = parser.parse_args()
    raiz = args.raiz.resolve()

    datos_dir = raiz / "datos"
    xlsx_candidatos = list(datos_dir.glob("*.xlsx"))
    if not xlsx_candidatos:
        raise FileNotFoundError(f"No se encontró ningún .xlsx en {datos_dir}")
    # Preferir el xlsx que contenga al menos una hoja de mediciones.
    hojas_esperadas = set(HOJAS.keys())
    xlsx_path = next(
        (p for p in xlsx_candidatos
         if hojas_esperadas & set(load_workbook(p, read_only=True).sheetnames)),
        xlsx_candidatos[0],
    )
    print(f"Leyendo: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    csv_dir = datos_dir / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)

    for hoja_nombre, csv_nombre in HOJAS.items():
        if hoja_nombre not in wb.sheetnames:
            print(f"  [SKIP] Hoja '{hoja_nombre}' no encontrada")
            continue
        ws = wb[hoja_nombre]
        encabezados, filas = extraer_hoja_ancho(ws)
        if not filas:
            print(f"  [SKIP] {hoja_nombre}: sin datos")
            continue
        out_path = csv_dir / csv_nombre
        with out_path.open("w", encoding="utf-8", newline="\n") as fh:
            writer = csv.writer(fh, lineterminator="\n")
            writer.writerow(encabezados)
            for fila in filas:
                writer.writerow(fila)
        n_semanas = len(encabezados) - 1
        print(f"  [OK] {csv_nombre}: {len(filas)} atributos, {n_semanas} semanas")

    print("Listo.")


if __name__ == "__main__":
    main()
